"""
Bot.py
Author: Jeremy King
Version: 0.01b
Change Log:
*****0.01b*****
Converted from JSON logging to xlsx format for easier data manipulation and the ability to append to files with minimal headache!
*****0.01a:*****
Built bot with PRAW and CSV to pull log and download whale post images
TODO: For continued post scraping add check to ensure post hasn't already been scrapped then skip if it has.
TODO: Add logging for posts access and keep database of posts containing 'whale' for future training and validation
TODO: Build function to call CNN built with pytorch
"""

import praw
from creds import r
import torch
import numpy
import os
import logging
import csv
import cython
import json
import urllib.request as request
import openpyxl as op
import collections

global lim
global rows
global log_created
lim = 20
sub = 'whales'


def bot_main(lim, sub):
    global rows
    global log_created
    log_created = 0
    reddit = r
    if os.path.isfile('log.xlsx'):
        print('Log exists, skipping!')
        log_created += 1
    else:
        make_xlxs()
    if log_created == 1:
        ms = op.load_workbook('log.xlsx')
        ws1 = ms['Logs']
        ws = ms.active
        rows = ws1.max_row

    subreddit = reddit.subreddit(sub).hot(limit=lim)
    for submission in subreddit:
        process_submission(submission)


def make_xlxs():
    global log_created
    print('Log not found, creating!')
    book = op.Workbook()
    ws = book.active
    log1 = book.create_sheet("Logs")
    log2 = book.create_sheet("Filtered Logs")

    book.save('log.xlsx')
    print('log file created successfully!')
    log_created += 1
    return


handler = logging.StreamHandler()
handler.setLevel(logging.DEBUG)
for logger_name in ("praw", "prawcore"):
    logger = logging.getLogger(logger_name)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)

# logging function
list_of_stuff = []
fields = ('title', 'url', 'id')


def process_submission(submission):
    global rows
    logs = 0
    #print(submission)
    if logs is not lim:
        to_dict = vars(submission)
        sub_dict = {field: to_dict[field] for field in fields}
        list_of_stuff.append(sub_dict)
        ms = op.load_workbook('log.xlsx')
        ws1 = ms['Logs']
        ws = ms.active
        post_id = sub_dict['id']
        post_title = sub_dict['title']
        url = sub_dict['url']
        #rows = int(ws.max_row)
        ws1.cell(column=1, row=rows, value=str(post_id))
        ws1.cell(column=2, row=rows, value=str(url))
        ws1.cell(column=3, row=rows, value=str(post_title))
        rows += 1
        #ws_title.append(post_title)
        #ws.append(post_id)
        logs += 1
    ms.save('log.xlsx')
    filter_whales()


def filter_whales():
    is_whale = "whale"
    main_sheet = op.load_workbook('log.xlsx')
    ws = main_sheet.active #ws = work_sheet
    log_rows = ws.max_row
    unfilter_logs = main_sheet['Logs']
    filter_logs = main_sheet['Filtered Logs']
    end_rows = ws.max_row
    cells = ws['A1':'A' + str(end_rows)]
    for c1 in cells:
        print(c1)

    '''
    for row in range(2, log_rows+1):
        for column in "A":
            cell_name = "{}{}".format(column, row)
            test = unfilter_logs[cell_name].value
            filter_logs.cell(column=1, row=rows, value=test)
            print(test)
            '''
    ms.save('log.xlsx')
    #unfilter_logs_value = unfilter_logs.cell(row=log_rows, column=1)
    #temp_dict = collections.defaultdict(unfilter_logs_value)
    #print(unfilter_logs_value)




def get_image():

    data = "aads"
    with open('filter.csv', 'r') as url_dict:
        read = csv.reader(url_dict)
        for row in read:
            filter_id = row[0]
            filter_url = row[1]
            whale_dict2 = {'id': row[0], 'url': row[1]}
            print(whale_dict2)
    end_doc = len(data)
    print(end_doc)
    print(url)
    filename = 0
    file_name =str(filter_id) + ".jpg"
    #path = os.path.join(pictures/, file_name)
    if filename is not end_doc:
        request.urlretrieve(url,file_name)
        filename +=1

#post = input("Enter number of posts to scrape! ")
#postint = int(post)
#sub = input("Enter subreddit without /r/ ")
bot_main(10, 'whales')
#filter_whales()
#get_image()

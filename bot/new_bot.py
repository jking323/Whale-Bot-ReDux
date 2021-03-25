"""
New_bot.py
Author: Jeremy King
Version 0.01a
*******0.01a*******
Initial
"""

import praw
from creds import r
import torch
import numpy
import os
import logging
import csv
import cython
import urllib.request as request
import openpyxl as op
import collections

#sub = 'whales'
def bot_main(lim, sub):
    
    log_created = 0
    reddit = r
    if os.path.isfile('log.xlsx'):
        print('Log exists, skipping!')
        log_created += 1
    else:
        make_xlxs()
    if log_created == 1:
        main_sheet = op.load_workbook('log.xlsx')
        unfiltered_work_sheet = main_sheet['Logs']
        working_sheet = main_sheet.active
        row_count = unfiltered_work_sheet
    subreddit = reddit.subreddit(sub).hot(limit=lim)
    for submission in subreddit:
        process_submission(submission)
    filter_whales()



def make_xlxs():
    print('Log not found, Creating!')
    book = op.workbook
    ws = book.active
    log1 = book.create_sheet("Logs")
    log2 = book.create_sheet("Filtered logs")

    book.save('log.xlsx')
    return

#global dict
post_data = []
data_fields = ('title','id','url')

def procss_submission(submission, lim):
    post_count = 0
    if post_count is not lim:
        write_to_dict = vars(submission)
        sub_dictonary = {field: write_to_dict[field] for field in fields}
        post_data.append(sub_dictonary)
        main_sheet = op.load_workbook('log.xlxs')
        working_sheet_unfiltered = main_sheet['Logs']
        active_working_sheet = main_sheet.active
        post_id = sub_dictonary['id']
        post_title = sub_dictonary['title']
        url = sub_dictonary['url']
        rows = int(active_working_sheet.max_row)
        active_working_sheet.cell(column=1, row=rows,value=str(post_id))
        active_working_sheet.cell(column=2, row=rows, value=str(url))
        active_working_sheet.cell(column=3, row=rows, value=str(post_title))
        logs += 1
    ms.save('log.xlsx')
    #filter_whales()

def filter_whales():
    is_whale = "whale"
    main_sheet = op.load_workbook('log.xlsx')
    working_sheet = main_sheet.active
    log_rows = working_sheet.max_row
    unfiltered_log = main_sheet['Logs']
    filtered_log = main_sheet['Filtered Logs']
    end_row = ws.max_row
    cells = unfiltered_log['A1':'A' + str(end_row)]
    for c1 in cells:
        print(cells)
bot_main(10, 'whales')

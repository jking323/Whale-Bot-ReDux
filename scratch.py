from openpyxl import Workbook as wb

book = wb()
ws = book.active

title1 = ws['A1']
title2 = ws['B1']
title1.value = 'id'
title2.value = 'url'

test_dict = {'post_id':'hvcae', 'url':'https://i.redd.it/q1ketsifrj661.jpg'}

test_id = test_dict['post_id']
url = test_dict['url']

a = ws['A2']
b = ws['B2']

a.value = test_id
b.value = url

book.save('test.xlsx')
print("sucess")